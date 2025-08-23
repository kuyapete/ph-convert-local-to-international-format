import pandas as pd
import openpyxl
from openpyxl import load_workbook

def convert_phone_numbers(file_path, output_path=None):
    """
    Convert mobile numbers in Excel file to country code format (63 prefix)
    and create a clean output without empty rows.
    
    Args:
        file_path (str): Path to the input Excel file
        output_path (str): Path for the output file (optional, defaults to input file)
    """
    
    # Load the workbook and select the active sheet
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Create a new worksheet for the cleaned data
    new_ws = wb.create_sheet("Converted_Data")
    
    # Copy header row if it exists
    if ws['A1'].value:
        new_ws['A1'] = ws['A1'].value
    
    # Track processed rows, empty rows, and invalid numbers
    processed_rows = 0
    empty_rows = 0
    invalid_numbers = 0
    new_row = 2  # Start from row 2 in new sheet
    
    # Process rows from 2 to 224680
    for row in range(2, 224681):  # 224681 because range is exclusive of the end
        cell = ws[f'A{row}']
        
        # Check if the cell is empty or None
        if cell.value is None or str(cell.value).strip() == '':
            empty_rows += 1
            continue
        
        # Convert the phone number
        phone_number = str(cell.value).strip()
        
        # Remove any non-digit characters (in case there are spaces or dashes)
        phone_number = ''.join(filter(str.isdigit, phone_number))
        
        # Convert local phone numbers to international format
        if phone_number.startswith('9'):
            # If starts with 9, add 63 prefix
            phone_number = '63' + phone_number
        elif phone_number.startswith('0'):
            # If starts with 0, add 63 prefix and remove the starting 0
            phone_number = '63' + phone_number[1:]
        
        # Validate Philippine phone number format
        # PH numbers should be exactly 12 digits starting with 63
        if phone_number.startswith('63') and len(phone_number) == 12:
            # Add to new worksheet
            new_ws[f'A{new_row}'] = phone_number
            new_row += 1
            processed_rows += 1
        else:
            # Skip invalid numbers
            invalid_numbers += 1
    
    # Remove the original sheet and rename the new one
    wb.remove(ws)
    new_ws.title = "Sheet1"
    
    # Save the file
    if output_path is None:
        output_path = file_path
    
    wb.save(output_path)
    
    print(f"Processing complete!")
    print(f"Converted {processed_rows} valid Philippine phone numbers to international format")
    print(f"Skipped {empty_rows} empty rows")
    print(f"Filtered out {invalid_numbers} invalid/non-Philippine numbers")
    print(f"File saved as: {output_path}")

# Example usage
if __name__ == "__main__":
    # Replace 'your_file.xlsx' with the actual path to your Excel file
    input_file = "PBET.xlsx"  # Change this to your file path
    output_file = "converted_phone_numbers.xlsx"  # Optional: specify output file name
    
    try:
        # Option 1: Overwrite the original file
        convert_phone_numbers(input_file)
        
        # Option 2: Save to a new file (uncomment the line below)
        # convert_phone_numbers(input_file, output_file)
        
    except FileNotFoundError:
        print(f"Error: File '{input_file}' not found. Please check the file path.")
    except Exception as e:
        print(f"An error occurred: {str(e)}")